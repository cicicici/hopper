from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import os
import sys
import datetime

from copy import copy, deepcopy
from openpyxl import load_workbook

from ..util.opt import Opt
from ..debug import log, dump

from .sheet import load_sheet, clear_sheet_cache


def load_book(filename, title_sheet_map, title_field_map):
    wb = load_workbook(filename)

    sheets = Opt()
    names = []

    for sheetname in wb.sheetnames:
        if sheetname in title_sheet_map:
            name = title_sheet_map[sheetname]
        else:
            name = sheetname

        sheets[name] = load_sheet(wb, sheetname, title_field_map)
        names.append({sheetname: name})

    log.trace(log.DC.STD, "Book: {}, sheets {}".format(filename, names))

    return Opt(wb=wb, sheets=sheets)

def clear_book_cache(book):
    for name, sheet in book.sheets.items():
        clear_sheet_cache(sheet)

